#!/usr/bin/env python3
"""
市町村決算状況 Excel → data.json 変換スクリプト
Usage: python extract_data.py <input.xlsx> [output.json]
"""

import sys
from pathlib import Path
import polars as pl
from multiprocessing import Pool
from datetime import datetime

import json


try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


def v(ws, coord):
    val = ws[coord].value
    if isinstance(val, datetime):
        return f"{val.month}-{val.day}"

    return val


def extract_sheet(ws):
    """1シート分のデータを辞書として返す。"""
    return {
        "年度": v(ws, 'D2'),
        "都道府県名": v(ws, 'BZ7'),
        "団体名": v(ws, 'CK7'),
        "都道府県コード": v(ws, 'BZ5'),
        "団体コード": v(ws, 'CK5'),
        "地方交付税種地": v(ws, 'DB5'),
        "市町村類型": v(ws, 'DB2'),
        "人口": {
            "令和2年国調": v(ws, 'AC2'),
            "平成27年国調": v(ws, 'AC3'),
            "増減率_percent": v(ws, 'AC4'),
            "住民基本台帳人口": {
                "令6_1_1": v(ws, 'AP4'),
                "令5_1_1": v(ws, 'AP5'),
                "増減率_percent": v(ws, 'AP6'),
            },
            "うち日本人": {
                "令6_1_1": v(ws, 'AX4'),
                "令5_1_1": v(ws, 'AX5'),
                "増減率_percent": v(ws, 'AX6'),
            },
        },
        "面積_km2": v(ws, 'AC5'),
        "人口密度_人_per_km2": v(ws, 'AC6'),
        "産業構造": {
            "第1次_令2国調": v(ws, 'BJ6'),
            "第1次_平27国調": v(ws, 'BQ6'),
            "第1次比率_percent": v(ws, 'BJ7'),
            "第2次_令2国調": v(ws, 'BJ8'),
            "第2次_平27国調": v(ws, 'BQ8'),
            "第2次比率_percent": v(ws, 'BJ9'),
            "第3次_令2国調": v(ws, 'BJ10'),
            "第3次_平27国調": v(ws, 'BQ10'),
            "第3次比率_percent": v(ws, 'BJ11'),
        },
        "歳入の状況_千円": {
            "地方税":                     _rev(ws, 'M11', 'U11', 'Y11', 'AG11'),
            "地方譲与税":                 _rev(ws, 'M12', 'U12', 'Y12', 'AG12'),
            "利子割交付金":               _rev(ws, 'M13', 'U13', 'Y13', 'AG13'),
            "配当割交付金":               _rev(ws, 'M14', 'U14', 'Y14', 'AG14'),
            "株式等譲渡所得割交付金":     _rev(ws, 'M15', 'U15', 'Y15', 'AG15'),
            "分離課税所得割交付金":       _rev(ws, 'M16', 'U16', 'Y16', 'AG16'),
            "地方消費税交付金":           _rev(ws, 'M17', 'U17', 'Y17', 'AG17'),
            "ゴルフ場利用税交付金":       _rev(ws, 'M18', 'U18', 'Y18', 'AG18'),
            "特別地方消費税交付金":       _rev(ws, 'M19', 'U19', 'Y19', 'AG19'),
            "自動車取得税交付金":         _rev(ws, 'M20', 'U20', 'Y20', 'AG20'),
            "軽油引取税交付金":           _rev(ws, 'M21', 'U21', 'Y21', 'AG21'),
            "自動車税環境性能割交付金":   _rev(ws, 'M22', 'U22', 'Y22', 'AG22'),
            "法人事業税交付金":           _rev(ws, 'M23', 'U23', 'Y23', 'AG23'),
            "地方特例交付金等": {
                **_rev(ws, 'M24', 'U24', 'Y24', 'AG24'),
                "内訳": {
                    "地方特例交付金":                           _rev(ws, 'M25', 'U25', 'Y25', 'AG25'),
                    "新型コロナウイルス感染症対策地方税減収補塡特別交付金": _rev(ws, 'M26', 'U26', 'Y26', 'AG26'),
                },
            },
            "地方交付税": {
                **_rev(ws, 'M27', 'U27', 'Y27', 'AG27'),
                "内訳": {
                    "普通交付税":       _rev(ws, 'M28', 'U28', 'Y28', 'AG28'),
                    "特別交付税":       _rev(ws, 'M29', 'U29', 'Y29', 'AG29'),
                    "震災復興特別交付税": _rev(ws, 'M30', 'U30', 'Y30', 'AG30'),
                },
            },
            "一般財源計":               _rev(ws, 'M31', 'U31', 'Y31', 'AG31'),
            "交通安全対策特別交付金":   _rev(ws, 'M32', 'U32', 'Y32', 'AG32'),
            "分担金負担金":             _rev(ws, 'M33', 'U33', 'Y33', 'AG33'),
            "使用料":                   _rev(ws, 'M34', 'U34', 'Y34', 'AG34'),
            "手数料":                   _rev(ws, 'M35', 'U35', 'Y35', 'AG35'),
            "国庫支出金":               _rev(ws, 'M36', 'U36', 'Y36', 'AG36'),
            "国有提供交付金":           _rev(ws, 'M37', 'U37', 'Y37', 'AG37'),
            "都道府県支出金":           _rev(ws, 'M39', 'U39', 'Y39', 'AG39'),
            "財産収入":                 _rev(ws, 'M40', 'U40', 'Y40', 'AG40'),
            "寄附金":                   _rev(ws, 'M41', 'U41', 'Y41', 'AG41'),
            "繰入金":                   _rev(ws, 'M42', 'U42', 'Y42', 'AG42'),
            "繰越金":                   _rev(ws, 'M43', 'U43', 'Y43', 'AG43'),
            "諸収入":                   _rev(ws, 'M44', 'U44', 'Y44', 'AG44'),
            "地方債": {
                **_rev(ws, 'M45', 'U45', 'Y45', 'AG45'),
                "うち減収補塡債特例分": v(ws, 'M46'),
                "うち臨時財政対策債":   v(ws, 'M47'),
            },
            "歳入合計": _rev(ws, 'M48', 'U48', 'Y48', 'AG48'),
        },
        "市町村税の状況_千円": {
            "普通税": {
                "合計": _tax3(ws, 'AW16', 'BF16', 'BJ16'),
                "法定普通税": {
                    "合計": _tax3(ws, 'AW17', 'BF17', 'BJ17'),
                    "市町村民税": {
                        **_tax3(ws, 'AW18', 'BF18', 'BJ18'),
                        "個人均等割": _tax2(ws, 'AW19', 'BF19'),
                        "所得割":     _tax2(ws, 'AW20', 'BF20'),
                        "法人均等割": _tax2(ws, 'AW21', 'BF21'),
                        "法人税割":   _tax2(ws, 'AW22', 'BF22'),
                    },
                    "固定資産税": {
                        **_tax2(ws, 'AW23', 'BF23'),
                        "うち純固定資産税": _tax2(ws, 'AW24', 'BF24'),
                    },
                    "軽自動車税":       _tax2(ws, 'AW25', 'BF25'),
                    "市町村たばこ税":   _tax2(ws, 'AW26', 'BF26'),
                    "鉱産税":           _tax2(ws, 'AW27', 'BF27'),
                    "特別土地保有税":   _tax2(ws, 'AW28', 'BF28'),
                },
                "法定外普通税": _tax2(ws, 'AW29', 'BF29'),
            },
            "目的税": {
                "合計": _tax2(ws, 'AW30', 'BF30'),
                "法定目的税": {
                    "入湯税":       _tax2(ws, 'AW32', 'BF32'),
                    "事業所税":     _tax2(ws, 'AW33', 'BF33'),
                    "都市計画税":   _tax2(ws, 'AW34', 'BF34'),
                    "水利地益税等": _tax2(ws, 'AW35', 'BF35'),
                },
                "法定外目的税": _tax2(ws, 'AW36', 'BF36'),
            },
            "旧法による税": _tax2(ws, 'AW37', 'BF37'),
            "合計": _tax3(ws, 'AW38', 'BF38', 'BJ38'),
        },
        "指定団体等の指定状況": {
            "旧新産":     v(ws, 'BV14') == '○',
            "旧工特":     v(ws, 'BV15') == '○',
            "低開発":     v(ws, 'BV16') == '○',
            "旧産炭":     v(ws, 'BV17') == '○',
            "山振":       v(ws, 'BV18') == '○',
            "過疎":       v(ws, 'BV19') == '○',
            "首都":       v(ws, 'BV20') == '○',
            "近畿":       v(ws, 'BV21') == '○',
            "中部":       v(ws, 'BV22') == '○',
            "財政健全化等": v(ws, 'BV23') == '○',
            "指数表選定": v(ws, 'BV24') == '○',
            "財源超過":   v(ws, 'BV25') == '○',
        },
        "収支状況_千円": {
            "歳入総額":             _yr2(ws, 'CO11', 'CY11'),
            "歳出総額":             _yr2(ws, 'CO12', 'CY12'),
            "歳入歳出差引":         _yr2(ws, 'CO13', 'CY13'),
            "翌年度に繰越すべき財源": _yr2(ws, 'CO14', 'CY14'),
            "実質収支":             _yr2(ws, 'CO15', 'CY15'),
            "単年度収支":           _yr2(ws, 'CO16', 'CY16'),
            "積立金":               _yr2(ws, 'CO17', 'CY17'),
            "繰上償還金":           _yr2(ws, 'CO18', 'CY18'),
            "積立金取崩し額":       _yr2(ws, 'CO19', 'CY19'),
            "実質単年度収支":       _yr2(ws, 'CO20', 'CY20'),
        },
        "性質別歳出の状況_千円": {
            "人件費": {
                **_exp(ws, 'M52', 'U52', 'Y52', 'AG52', 'AQ52'),
                "うち職員給": {"決算額": v(ws, 'M53'), "充当一般財源等": v(ws, 'Y53')},
            },
            "扶助費":       _exp(ws, 'M54', 'U54', 'Y54', 'AG54', 'AQ54'),
            "公債費": {
                **_exp(ws, 'M55', 'U55', 'Y55', 'AG55', 'AQ55'),
                "元利償還金": {
                    "元金":         _exp(ws, 'M56', None, 'Y56', 'AG56', 'AQ56'),
                    "利子":         _exp(ws, 'M57', None, 'Y57', 'AG57', 'AQ57'),
                    "一時借入金利子": _exp(ws, 'M58', None, 'Y58', 'AG58', 'AQ58'),
                },
            },
            "義務的経費計":   _exp(ws, 'M59', 'U59', 'Y59', 'AG59', 'AQ59'),
            "物件費":         _exp(ws, 'M60', 'U60', 'Y60', 'AG60', 'AQ60'),
            "維持補修費":     _exp(ws, 'M61', 'U61', 'Y61', 'AG61', 'AQ61'),
            "補助費等": {
                **_exp(ws, 'M62', 'U62', 'Y62', 'AG62', 'AQ62'),
                "うち一部事務組合負担金": _exp(ws, 'M63', None, 'Y63', 'AG63', 'AQ63'),
            },
            "繰出金":         _exp(ws, 'M64', 'U64', 'Y64', 'AG64', 'AQ64'),
            "積立金":         {"決算額": v(ws, 'M65'), "構成比": v(ws, 'U65'), "充当一般財源等": v(ws, 'Y65')},
            "投資出資金貸付金": _exp(ws, 'M66', 'U66', 'Y66', 'AG66', 'AQ66'),
            "前年度繰上充用金": {"決算額": v(ws, 'M67'), "構成比": v(ws, 'U67'), "充当一般財源等": v(ws, 'Y67')},
            "投資的経費": {
                "決算額": v(ws, 'M68'), "構成比": v(ws, 'U68'), "充当一般財源等": v(ws, 'Y68'),
                "うち人件費": {"決算額": v(ws, 'M69'), "充当一般財源等": v(ws, 'Y69')},
                "普通建設事業費": {
                    "決算額": v(ws, 'M70'), "構成比": v(ws, 'U70'), "充当一般財源等": v(ws, 'Y70'),
                    "うち補助": {"決算額": v(ws, 'M71'), "構成比": v(ws, 'U71'), "充当一般財源等": v(ws, 'Y71')},
                    "うち単独": {"決算額": v(ws, 'M72'), "構成比": v(ws, 'U72'), "充当一般財源等": v(ws, 'Y72')},
                },
                "災害復旧事業費": {"決算額": v(ws, 'M73'), "構成比": v(ws, 'U73'), "充当一般財源等": v(ws, 'Y73')},
                "失業対策事業費": {"決算額": v(ws, 'M74'), "構成比": v(ws, 'U74'), "充当一般財源等": v(ws, 'Y74')},
            },
            "歳出合計": {"決算額": v(ws, 'M75'), "構成比": v(ws, 'U75'), "充当一般財源等": v(ws, 'Y75')},
            "経常経費充当一般財源等計_千円": v(ws, 'AH69'),
            "経常収支比率_percent": v(ws, 'AK71'),
            "歳入一般財源等_千円": v(ws, 'AH75'),
        },
        "目的別歳出の状況_千円": {
            "議会費":       _mok(ws, 'BF53', 'BM53', 'BQ53', 'BZ53'),
            "総務費":       _mok(ws, 'BF54', 'BM54', 'BQ54', 'BZ54'),
            "民生費":       _mok(ws, 'BF55', 'BM55', 'BQ55', 'BZ55'),
            "衛生費":       _mok(ws, 'BF56', 'BM56', 'BQ56', 'BZ56'),
            "労働費":       _mok(ws, 'BF57', 'BM57', 'BQ57', 'BZ57'),
            "農林水産業費": _mok(ws, 'BF58', 'BM58', 'BQ58', 'BZ58'),
            "商工費":       _mok(ws, 'BF59', 'BM59', 'BQ59', 'BZ59'),
            "土木費":       _mok(ws, 'BF60', 'BM60', 'BQ60', 'BZ60'),
            "消防費":       _mok(ws, 'BF61', 'BM61', 'BQ61', 'BZ61'),
            "教育費":       _mok(ws, 'BF62', 'BM62', 'BQ62', 'BZ62'),
            "災害復旧費":   _mok(ws, 'BF63', 'BM63', 'BQ63', 'BZ63'),
            "公債費":       _mok(ws, 'BF64', 'BM64', 'BQ64', 'BZ64'),
            "諸支出金":     _mok(ws, 'BF65', 'BM65', 'BQ65', 'BZ65'),
            "前年度繰上充用金": _mok(ws, 'BF66', 'BM66', 'BQ66', 'BZ66'),
            "歳出合計":     _mok(ws, 'BF67', 'BM67', 'BQ67', 'BZ67'),
            "公営事業等への繰出合計_千円": v(ws, 'BF69'),
        },
        "財政指標": {
            "基準財政収入額_千円":  _yr2(ws, 'CS51', 'DA51'),
            "基準財政需要額_千円":  _yr2(ws, 'CS52', 'DA52'),
            "標準税収入額等_千円":  _yr2(ws, 'CS53', 'DA53'),
            "標準財政規模_千円":    _yr2(ws, 'CS54', 'DA54'),
            "財政力指数":           _yr2(ws, 'CS55', 'DA55'),
            "実質収支比率_percent": _yr2(ws, 'CS56', 'DA56'),
            "公債費負担比率_percent": _yr2(ws, 'CS57', 'DA57'),
            "健全化判断比率": {
                "実質赤字比率_percent":       _yr2(ws, 'CS58', 'DA58'),
                "連結実質赤字比率_percent":   _yr2(ws, 'CS59', 'DA59'),
                "実質公債費比率_percent":     _yr2(ws, 'CS60', 'DA60'),
                "将来負担比率_percent":       _yr2(ws, 'CS61', 'DA61'),
            },
        },
        "積立金現在高_千円": {
            "財調":     _yr2(ws, 'CS62', 'DA62'),
            "減債":     _yr2(ws, 'CS63', 'DA63'),
            "特定目的": _yr2(ws, 'CS64', 'DA64'),
        },
        "地方債現在高_千円": _yr2(ws, 'CS65', 'DA65'),
        "債務負担行為額支出予定額_千円": {
            "物件等購入": _yr2(ws, 'CS66', 'DA66'),
            "保証補償":   _yr2(ws, 'CS67', 'DA67'),
            "その他":     _yr2(ws, 'CS68', 'DA68'),
        },
        "一部事務組合加入の状況": {
            "議員公務災害":   v(ws, 'BV32') == '○',
            "非常勤公務災害": v(ws, 'BV33') == '○',
            "退職手当":       v(ws, 'BV34') == '○',
            "事務機共同":     v(ws, 'BV35') == '○',
            "税務事務":       v(ws, 'BV36') == '○',
            "老人福祉":       v(ws, 'BV37') == '○',
            "伝染病":         v(ws, 'BV38') == '○',
            "し尿処理":       v(ws, 'CC32') == '○',
            "ごみ処理":       v(ws, 'CC33') == '○',
            "火葬場":         v(ws, 'CC34') == '○',
            "常備消防":       v(ws, 'CC35') == '○',
            "小学校":         v(ws, 'CC36') == '○',
            "中学校":         v(ws, 'CC37') == '○',
            "その他":         v(ws, 'CC38') == '○',
        },
        "職員給与の状況": {
            "一般職員等": {
                "一般職員":     _staff(ws, 'CO23', 'CT23', 'CZ23'),
                "うち消防職員": _staff(ws, 'CO24', 'CT24', 'CZ24'),
                "うち技能労務員": _staff(ws, 'CO25', 'CT25', 'CZ25'),
            },
            "教育公務員": _staff(ws, 'CO26', 'CT26', 'CZ26'),
            "臨時職員":   _staff(ws, 'CO27', 'CT27', 'CZ27'),
            "合計":       _staff(ws, 'CO28', 'CT28', 'CZ28'),
            "ラスパイレス指数": v(ws, 'CO29'),
            "特別職等": {
                "市区町村長":   _sp(ws, 'CO32', 'CT32', 'CZ32'),
                "副市区町村長": _sp(ws, 'CO33', 'CT33', 'CZ33'),
                "教育長":       _sp(ws, 'CO34', 'CT34', 'CZ34'),
                "議会議長":     _sp(ws, 'CO35', 'CT35', 'CZ35'),
                "議会副議長":   _sp(ws, 'CO36', 'CT36', 'CZ36'),
                "議会議員":     _sp(ws, 'CO37', 'CT37', 'CZ37'),
            },
        },
        "会計の状況": {
            "国民健康保険事業": {
                "実質収支_千円":    v(ws, 'BZ69'),
                "再差引収支_千円":  v(ws, 'BZ70'),
                "加入世帯数_世帯":  v(ws, 'BZ71'),
                "被保険者数_人":    v(ws, 'BZ72'),
                "被保険者1人当り": {
                    "保険税料収入額_円": v(ws, 'BZ73'),
                    "国庫支出金_円":     v(ws, 'BZ74'),
                    "保険給付費_円":     v(ws, 'BZ75'),
                },
            },
            "公営事業等への繰出": {
                "下水道_千円":       v(ws, 'BF70'),
                "交通_千円":         v(ws, 'BF71'),
                "病院_千円":         v(ws, 'BF72'),
                "上水道_千円":       v(ws, 'BF73'),
                "国民健康保険_千円": v(ws, 'BF74'),
                "その他_千円":       v(ws, 'BF75'),
            },
        },
        "収益事業収入_千円":      _yr2(ws, 'CS70', 'DA70'),
        "土地開発基金現在高_千円": _yr2(ws, 'CS71', 'DA71'),
        "徴収率_percent": {
            "市町村民税": {
                "令和5年度": {"現年": v(ws, 'CS74'), "計": v(ws, 'DA74')},
                "令和4年度": {"現年": v(ws, 'DE74')},
            },
            "純固定資産税": {
                "令和5年度": {"現年": v(ws, 'CS75'), "計": v(ws, 'DA75')},
                "令和4年度": {"現年": v(ws, 'DE75')},
            },
            "合計": {
                "令和5年度": {"現年": v(ws, 'CS72'), "計": v(ws, 'CW72')},
                "令和4年度": {"現年": v(ws, 'DA72'), "計": v(ws, 'DE72')},
            },
        },
    }


# ── ヘルパー関数 ────────────────────────────────────────────

def _rev(ws, dec, comp, josei, josei_comp):
    """歳入行 (決算額, 構成比, 経常一般財源等, 経常一般財源等構成比)"""
    return {
        "決算額": v(ws, dec),
        "構成比": v(ws, comp),
        "経常一般財源等": v(ws, josei),
        "経常一般財源等構成比": v(ws, josei_comp),
    }


def _tax2(ws, col_収入, col_構成比):
    return {"収入済額": v(ws, col_収入), "構成比": v(ws, col_構成比)}


def _tax3(ws, col_収入, col_構成比, col_超過):
    return {"収入済額": v(ws, col_収入), "構成比": v(ws, col_構成比), "超過課税分": v(ws, col_超過)}


def _yr2(ws, col5, col4):
    return {"令和5年度": v(ws, col5), "令和4年度": v(ws, col4)}


def _exp(ws, dec, comp, juto, keijo_juto, keijo_ratio):
    d = {"決算額": v(ws, dec), "充当一般財源等": v(ws, juto)}
    if comp:
        d["構成比"] = v(ws, comp)
    if keijo_juto:
        d["経常経費充当一般財源等"] = v(ws, keijo_juto)
    if keijo_ratio:
        d["経常収支比率"] = v(ws, keijo_ratio)
    return d


def _mok(ws, dec, comp, kensetsu, juto):
    return {
        "決算額": v(ws, dec),
        "構成比": v(ws, comp),
        "普通建設事業費": v(ws, kensetsu),
        "充当一般財源等": v(ws, juto),
    }


def _staff(ws, ninzu, gakko, hitori):
    return {
        "職員数_人": v(ws, ninzu),
        "給料月額_百円": v(ws, gakko),
        "一人当たり平均給料月額_百円": v(ws, hitori),
    }


def _sp(ws, teisu, tekiyo, hitori):
    return {
        "定数": v(ws, teisu),
        "適用開始年月日": v(ws, tekiyo),
        "一人当たり平均給料報酬月額_百円": v(ws, hitori),
    }


# ── CSVに変換する処理 ──────────────────────────────────────────────
def flatten(nested: dict, sep: str = ".") -> pl.DataFrame:
    """Recursively flatten a nested dict."""
    out = {}

    def _flatten(obj, prefix=""):
        if obj == "-" or obj == "*":
            obj = None

        if isinstance(obj, dict):
            for k, v in obj.items():
                _flatten(v, f"{prefix}{k}{sep}" if prefix else f"{k}{sep}")
        else:
            out[prefix.rstrip(sep)] = obj

    _flatten(nested)
    return out

def add_cols(df: pl.DataFrame):
    # 市町村税について、歳入構成比追加
    stack1 = []
    stack2 = []
    for c in df.columns:
        check = '市町村税の状況_千円' in i and ('超過課税分' in i or '収入済額' in i)
        if check:
            stack1.append(pl.col(c).alias(f"{c}歳入合計構成比"))
            stack2.append(pl.col(c).alias(f"{c}歳入合計経常一般財源等構成比"))

    print(f"col count: {len(df.columns)}")
    df = df.with_columns(pl.col(stack1)/pl.col('歳入の状況_千円.歳入合計.決算額'))
    print(f"col count2: {len(df.columns)}")
    df = df.with_columns(pl.col(stack2)/pl.col('歳入の状況_千円.歳入合計.経常一般財源等'))
    print(f"col count3: {len(df.columns)}")

    return df

# ── メイン処理 ──────────────────────────────────────────────
def main(xlsx_path: Path):
    print(f"読み込み中: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = [s for s in wb.sheetnames if s != '目次']
    print(f"対象シート数: {len(sheets)}")

    errors = []
    stack = []
    for i, name in enumerate(sheets, 1):
        try:
            data = flatten(extract_sheet(wb[name]))
            _ = json.dumps(data)
            stack.append(data)
            print(f"  [{i:3d}/{len(sheets)}] {name} … OK")
        except Exception as e:
            errors.append((name, str(e)))
            print(data)
            print(f"  [{i:3d}/{len(sheets)}] {name} … ERROR: {e}")
        
    print(f"\n完了: {len(stack)} 件")
    if errors:
        print(f"エラー {len(errors)} 件:")
        for name, msg in errors:
            print(f"  {name}: {msg}")
    
    return stack

if __name__ == '__main__':
    input_xlsx = Path(sys.argv[1])
    print(f"決算データのあるディレクトリ: {input_xlsx}")
    output_path = Path(sys.argv[2])
    print(f"出力するディレクトリ: {output_path}")

    pool = Pool(4)
    all_data = []
    for i in pool.map(main, input_xlsx.iterdir()):
        all_data.extend(i)

    s = json.dumps(all_data, ensure_ascii=False, indent=2)

    fp = open("./a.json", "w+")
    fp.write(s)
    fp.close()

    df = pl.DataFrame(all_data, infer_schema_length=len(all_data))
    df = df.filter(pl.col("年度").is_not_null())
    df = add_cols(df)
    df.write_csv(Path(output_path).with_name("r5_kessan_data.csv"))
