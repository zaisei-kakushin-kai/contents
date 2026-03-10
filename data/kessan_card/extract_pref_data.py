#!/usr/bin/env python3
"""
都道府県決算状況 Excel → JSON/msgpack 変換スクリプト
Usage: python extract_pref_data.py <都道府県.xlsx> [output.json]
"""

import sys
import json
from pathlib import Path
from datetime import datetime
import msgpack
import polars as pl

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl が必要です: pip install openpyxl")
    sys.exit(1)


PERF_LIST = [
    "北海道",
    "青森県",
    "岩手県",
    "宮城県",
    "秋田県",
    "山形県",
    "福島県",
    "茨城県",
    "栃木県",
    "群馬県",
    "埼玉県",
    "千葉県",
    "東京都",
    "神奈川県",
    "新潟県",
    "富山県",
    "石川県",
    "福井県",
    "山梨県",
    "長野県",
    "岐阜県",
    "静岡県",
    "愛知県",
    "三重県",
    "滋賀県",
    "京都府",
    "大阪府",
    "兵庫県",
    "奈良県",
    "和歌山県",
    "鳥取県",
    "島根県",
    "岡山県",
    "広島県",
    "山口県",
    "徳島県",
    "香川県",
    "愛媛県",
    "高知県",
    "福岡県",
    "佐賀県",
    "長崎県",
    "熊本県",
    "大分県",
    "宮崎県",
    "鹿児島県",
    "沖縄県"
]

# ── セル取得ヘルパー ──────────────────────────────────────────


def v(ws, coord):
    val = ws[coord].value
    if isinstance(val, datetime):
        return f"{val.month}-{val.day}"
    return val


# ── データ構造ヘルパー ────────────────────────────────────────


def _rev(ws, dec, comp, keijo, keijo_comp):
    """歳入行 (決算額, 構成比, 経常一般財源等, 経常一般財源等構成比)"""
    return {
        "決算額": v(ws, dec),
        "構成比": v(ws, comp),
        "経常一般財源等": v(ws, keijo),
        "経常一般財源等構成比": v(ws, keijo_comp),
    }


def _tax3(ws, col_収入, col_構成比, col_超過):
    return {
        "収入済額": v(ws, col_収入),
        "構成比": v(ws, col_構成比),
        "超過課税分": v(ws, col_超過),
    }


def _yr2(ws, col5, col4):
    return {"令和5年度": v(ws, col5), "令和4年度": v(ws, col4)}


def _exp(ws, dec, comp, juto, keijo_juto=None, keijo_ratio=None):
    """性質別歳出行"""
    d = {"決算額": v(ws, dec), "充当一般財源等": v(ws, juto)}
    if comp:
        d["構成比"] = v(ws, comp)
    if keijo_juto:
        d["経常経費充当一般財源等"] = v(ws, keijo_juto)
    if keijo_ratio:
        d["経常収支比率"] = v(ws, keijo_ratio)
    return d


def _mok(ws, dec, comp, kensetsu, juto):
    """目的別歳出行"""
    return {
        "決算額": v(ws, dec),
        "構成比": v(ws, comp),
        "普通建設事業費": v(ws, kensetsu),
        "充当一般財源等": v(ws, juto),
    }


def _staff(ws, ninzu, kyuryo, hitori):
    return {
        "職員数_人": v(ws, ninzu),
        "給料月額_百円": v(ws, kyuryo),
        "一人当たり平均給料月額_百円": v(ws, hitori),
    }


def _sp(ws, teisu, tekiyo, hitori):
    return {
        "定数": v(ws, teisu),
        "適用開始年月日": v(ws, tekiyo),
        "一人当たり平均給料報酬月額_百円": v(ws, hitori),
    }


# ── シート抽出 ────────────────────────────────────────────────


def extract_sheet(ws):
    """1シート分（1都道府県）のデータを辞書として返す。"""
    return {
        "年度": v(ws, 'D2'),
        "都道府県コード": v(ws, 'CO2'),
        "都道府県名": v(ws, 'CQ2'),
        "人口": {
            "令和2年国調": v(ws, 'AH2'),
            "平成27年国調": v(ws, 'AH3'),
            "増減率_percent": v(ws, 'AH4'),
            "住民基本台帳人口": {
                "令6_1_1": v(ws, 'AW4'),
                "令5_1_1": v(ws, 'AW5'),
                "増減率_percent": v(ws, 'AW6'),
            },
            "うち日本人": {
                "令6_1_1": v(ws, 'BF4'),
                "令5_1_1": v(ws, 'BF5'),
                "増減率_percent": v(ws, 'BF6'),
            },
        },
        "面積_km2": v(ws, 'AH5'),
        "人口密度_人_per_km2": v(ws, 'AH6'),
        "収支状況_千円": {
            "歳入総額":               _yr2(ws, 'CQ5',  'DA5'),
            "歳出総額":               _yr2(ws, 'CQ6',  'DA6'),
            "歳入歳出差引":           _yr2(ws, 'CQ7',  'DA7'),
            "翌年度に繰越すべき財源": _yr2(ws, 'CQ8',  'DA8'),
            "実質収支":               _yr2(ws, 'CQ9',  'DA9'),
            "単年度収支":             _yr2(ws, 'CQ10', 'DA10'),
            "積立金":                 _yr2(ws, 'CQ11', 'DA11'),
            "繰上償還金":             _yr2(ws, 'CQ12', 'DA12'),
            "積立金取崩し額":         _yr2(ws, 'CQ13', 'DA13'),
            "実質単年度収支":         _yr2(ws, 'CQ14', 'DA14'),
        },
        "歳入の状況_千円": {
            "地方税":                 _rev(ws, 'M11', 'U11', 'Z11', 'AH11'),
            "地方譲与税": {
                **_rev(ws, 'M12', 'U12', 'Z12', 'AH12'),
                "内訳": {
                    "地方揮発油譲与税":     _rev(ws, 'M13', 'U13', 'Z13', 'AH13'),
                    "特別とん譲与税":       _rev(ws, 'M14', 'U14', 'Z14', 'AH14'),
                    "石油ガス譲与税":       _rev(ws, 'M15', 'U15', 'Z15', 'AH15'),
                    "自動車重量譲与税":     _rev(ws, 'M16', 'U16', 'Z16', 'AH16'),
                    "航空機燃料譲与税":     _rev(ws, 'M17', 'U17', 'Z17', 'AH17'),
                    "森林環境譲与税":       _rev(ws, 'M18', 'U18', 'Z18', 'AH18'),
                    "特別法人事業譲与税":   _rev(ws, 'M19', 'U19', 'Z19', 'AH19'),
                },
            },
            "市町村たばこ税都道府県交付金": _rev(ws, 'M20', 'U20', 'Z20', 'AH20'),
            "地方特例交付金等": {
                **_rev(ws, 'M21', 'U21', 'Z21', 'AH21'),
                "内訳": {
                    "地方特例交付金":                           _rev(ws, 'M22', 'U22', 'Z22', 'AH22'),
                    "新型コロナウイルス感染症対策地方税減収補塡特別交付金": _rev(ws, 'M23', 'U23', 'Z23', 'AH23'),
                },
            },
            "地方交付税": {
                **_rev(ws, 'M24', 'U24', 'Z24', 'AH24'),
                "内訳": {
                    "普通交付税":           _rev(ws, 'M25', 'U25', 'Z25', 'AH25'),
                    "特別交付税":           _rev(ws, 'M26', 'U26', 'Z26', 'AH26'),
                    "震災復興特別交付税":   _rev(ws, 'M27', 'U27', 'Z27', 'AH27'),
                },
            },
            "一般財源計":                   _rev(ws, 'M28', 'U28', 'Z28', 'AH28'),
            "交通安全対策特別交付金":       _rev(ws, 'M29', 'U29', 'Z29', 'AH29'),
            "分担金・負担金":               _rev(ws, 'M30', 'U30', 'Z30', 'AH30'),
            "使用料":                       _rev(ws, 'M31', 'U31', 'Z31', 'AH31'),
            "手数料":                       _rev(ws, 'M32', 'U32', 'Z32', 'AH32'),
            "国庫支出金":                   _rev(ws, 'M33', 'U33', 'Z33', 'AH33'),
            "国有提供交付金":               _rev(ws, 'M34', 'U34', 'Z34', 'AH34'),
            "財産収入":                     _rev(ws, 'M35', 'U35', 'Z35', 'AH35'),
            "寄附金":                       _rev(ws, 'M36', 'U36', 'Z36', 'AH36'),
            "繰入金":                       _rev(ws, 'M37', 'U37', 'Z37', 'AH37'),
            "繰越金":                       _rev(ws, 'M38', 'U38', 'Z38', 'AH38'),
            "諸収入":                       _rev(ws, 'M39', 'U39', 'Z39', 'AH39'),
            "地方債": {
                **_rev(ws, 'M40', 'U40', 'Z40', 'AH40'),
                "うち減収補塡債特例分": v(ws, 'M41'),
                "うち臨時財政対策債":   v(ws, 'M42'),
            },
            "歳入合計": _rev(ws, 'M43', 'U43', 'Z43', 'AH43'),
        },
        "道府県税の状況_千円": {
            "普通税": {
                **_tax3(ws, 'BD11', 'BK11', 'BP11'),
                "法定普通税": {
                    **_tax3(ws, 'BD12', 'BK12', 'BP12'),
                    "道府県民税": {
                        **_tax3(ws, 'BD13', 'BK13', 'BP13'),
                        "個人均等割":       _tax3(ws, 'BD14', 'BK14', 'BP14'),
                        "所得割":           _tax3(ws, 'BD15', 'BK15', 'BP15'),
                        "法人均等割":       _tax3(ws, 'BD16', 'BK16', 'BP16'),
                        "法人税割":         _tax3(ws, 'BD17', 'BK17', 'BP17'),
                        "利子割":           _tax3(ws, 'BD18', 'BK18', 'BP18'),
                        "配当割":           _tax3(ws, 'BD19', 'BK19', 'BP19'),
                        "株式等譲渡所得割": _tax3(ws, 'BD20', 'BK20', 'BP20'),
                    },
                    "事業税": {
                        **_tax3(ws, 'BD21', 'BK21', 'BP21'),
                        "内訳": {
                            "個人分": _tax3(ws, 'BD22', 'BK22', 'BP22'),
                            "法人分": _tax3(ws, 'BD23', 'BK23', 'BP23'),
                        },
                    },
                    "地方消費税":       _tax3(ws, 'BD24', 'BK24', 'BP24'),
                    "不動産取得税":     _tax3(ws, 'BD25', 'BK25', 'BP25'),
                    "道府県たばこ税":   _tax3(ws, 'BD26', 'BK26', 'BP26'),
                    "ゴルフ場利用税":   _tax3(ws, 'BD27', 'BK27', 'BP27'),
                    "軽油引取税":       _tax3(ws, 'BD28', 'BK28', 'BP28'),
                    "自動車税":         _tax3(ws, 'BD29', 'BK29', 'BP29'),
                    "鉱区税":           _tax3(ws, 'BD30', 'BK30', 'BP30'),
                    "固定資産税特例":   _tax3(ws, 'BD31', 'BK31', 'BP31'),
                },
                "法定外普通税": _tax3(ws, 'BD32', 'BK32', 'BP32'),
            },
            "目的税": {
                **_tax3(ws, 'BD33', 'BK33', 'BP33'),
                "法定目的税": {
                    **_tax3(ws, 'BD34', 'BK34', 'BP34'),
                    "狩猟税": _tax3(ws, 'BD35', 'BK35', 'BP35'),
                },
                "法定外目的税": _tax3(ws, 'BD36', 'BK36', 'BP36'),
            },
            "旧法による税": _tax3(ws, 'BD37', 'BK37', 'BP37'),
            "合計": _tax3(ws, 'BD43', 'BK43', 'BP43'),
        },
        "性質別歳出の状況_千円": {
            "義務的経費計": {
                **_exp(ws, 'M48', 'U48', 'Z48', 'AH48', 'AP48'),
                "人件費": {
                    **_exp(ws, 'M49', 'U49', 'Z49', 'AH49', 'AP49'),
                    "うち職員給": _exp(ws, 'M50', 'U50', 'Z50', 'AH50', 'AP50'),
                },
                "扶助費": _exp(ws, 'M51', 'U51', 'Z51', 'AH51', 'AP51'),
                "公債費": {
                    **_exp(ws, 'M52', 'U52', 'Z52', 'AH52', 'AP52'),
                    "元利償還金": {
                        "元金":           _exp(ws, 'M53', 'U53', 'Z53', 'AH53', 'AP53'),
                        "利子":           _exp(ws, 'M54', 'U54', 'Z54', 'AH54', 'AP54'),
                        "一時借入金利子": _exp(ws, 'M55', 'U55', 'Z55', 'AH55', 'AP55'),
                    },
                },
            },
            "その他の経費": {
                **_exp(ws, 'M56', 'U56', 'Z56', 'AH56', 'AP56'),
                "物件費":         _exp(ws, 'M57', 'U57', 'Z57', 'AH57', 'AP57'),
                "維持補修費":     _exp(ws, 'M58', 'U58', 'Z58', 'AH58', 'AP58'),
                "補助費等":       _exp(ws, 'M59', 'U59', 'Z59', 'AH59', 'AP59'),
                "繰出金":         _exp(ws, 'M60', 'U60', 'Z60', 'AH60', 'AP60'),
                "積立金":         {"決算額": v(ws, 'M61'), "構成比": v(ws, 'U61'), "充当一般財源等": v(ws, 'Z61')},
                "投資及び出資金": {"決算額": v(ws, 'M62'), "構成比": v(ws, 'U62'), "充当一般財源等": v(ws, 'Z62')},
                "貸付金":         {"決算額": v(ws, 'M63'), "構成比": v(ws, 'U63'), "充当一般財源等": v(ws, 'Z63')},
                "前年度繰上充用金": {"決算額": v(ws, 'M64'), "構成比": v(ws, 'U64'), "充当一般財源等": v(ws, 'Z64')},
            },
            "投資的経費計": {
                "決算額": v(ws, 'M65'), "構成比": v(ws, 'U65'), "充当一般財源等": v(ws, 'Z65'),
                "うち人件費": {"決算額": v(ws, 'M66'), "構成比": v(ws, 'U66'), "充当一般財源等": v(ws, 'Z66')},
                "普通建設事業費": {
                    "決算額": v(ws, 'M67'), "構成比": v(ws, 'U67'), "充当一般財源等": v(ws, 'Z67'),
                    "うち補助": {"決算額": v(ws, 'M68'), "構成比": v(ws, 'U68'), "充当一般財源等": v(ws, 'Z68')},
                    "うち単独": {"決算額": v(ws, 'M69'), "構成比": v(ws, 'U69'), "充当一般財源等": v(ws, 'Z69')},
                },
                "災害復旧事業費": {"決算額": v(ws, 'M70'), "構成比": v(ws, 'U70'), "充当一般財源等": v(ws, 'Z70')},
                "失業対策事業費": {"決算額": v(ws, 'M71'), "構成比": v(ws, 'U71'), "充当一般財源等": v(ws, 'Z71')},
            },
            "歳出合計": {"決算額": v(ws, 'M72'), "構成比": v(ws, 'U72'), "充当一般財源等": v(ws, 'Z72')},
            "経常経費充当一般財源等計_千円": v(ws, 'AH66'),
            "経常収支比率_percent": v(ws, 'AO68'),
            "経常収支比率_臨財債等除_percent": v(ws, 'AO69'),
            "歳入一般財源等_千円": v(ws, 'AH72'),
        },
        "目的別歳出の状況_千円": {
            "議会費":           _mok(ws, 'BD48', 'BK48', 'BP48', 'BW48'),
            "総務費":           _mok(ws, 'BD49', 'BK49', 'BP49', 'BW49'),
            "民生費":           _mok(ws, 'BD50', 'BK50', 'BP50', 'BW50'),
            "衛生費":           _mok(ws, 'BD51', 'BK51', 'BP51', 'BW51'),
            "労働費":           _mok(ws, 'BD52', 'BK52', 'BP52', 'BW52'),
            "農林水産業費":     _mok(ws, 'BD53', 'BK53', 'BP53', 'BW53'),
            "商工費":           _mok(ws, 'BD54', 'BK54', 'BP54', 'BW54'),
            "土木費":           _mok(ws, 'BD55', 'BK55', 'BP55', 'BW55'),
            "警察費":           _mok(ws, 'BD56', 'BK56', 'BP56', 'BW56'),
            "消防費":           _mok(ws, 'BD57', 'BK57', 'BP57', 'BW57'),
            "教育費":           _mok(ws, 'BD58', 'BK58', 'BP58', 'BW58'),
            "災害復旧費":       _mok(ws, 'BD59', 'BK59', 'BP59', 'BW59'),
            "公債費":           _mok(ws, 'BD60', 'BK60', 'BP60', 'BW60'),
            "諸支出金":         _mok(ws, 'BD61', 'BK61', 'BP61', 'BW61'),
            "前年度繰上充用金": _mok(ws, 'BD62', 'BK62', 'BP62', 'BW62'),
            "利子割交付金":                 _mok(ws, 'BD63', 'BK63', 'BP63', 'BW63'),
            "配当割交付金":                 _mok(ws, 'BD64', 'BK64', 'BP64', 'BW64'),
            "株式等譲渡所得割交付金":       _mok(ws, 'BD65', 'BK65', 'BP65', 'BW65'),
            "分離課税所得割交付金":         _mok(ws, 'BD66', 'BK66', 'BP66', 'BW66'),
            "地方消費税交付金":             _mok(ws, 'BD67', 'BK67', 'BP67', 'BW67'),
            "ゴルフ場利用税交付金":         _mok(ws, 'BD68', 'BK68', 'BP68', 'BW68'),
            "特別地方消費税交付金":         _mok(ws, 'BD69', 'BK69', 'BP69', 'BW69'),
            "自動車取得税交付金":           _mok(ws, 'BD70', 'BK70', 'BP70', 'BW70'),
            "軽油引取税交付金":             _mok(ws, 'BD71', 'BK71', 'BP71', 'BW71'),
            "自動車税環境性能割交付金":     _mok(ws, 'BD72', 'BK72', 'BP72', 'BW72'),
            "法人事業税交付金":             _mok(ws, 'BD73', 'BK73', 'BP73', 'BW73'),
            "特別区財政調整交付金":         _mok(ws, 'BD74', 'BK74', 'BP74', 'BW74'),
            "歳出合計":         _mok(ws, 'BD75', 'BK75', 'BP75', 'BW75'),
        },
        "財政指標": {
            "基準財政収入額_千円":    _yr2(ws, 'CQ34', 'DA34'),
            "基準財政需要額_千円":    _yr2(ws, 'CQ35', 'DA35'),
            "標準税収入額等_千円":    _yr2(ws, 'CQ36', 'DA36'),
            "標準財政規模_千円":      _yr2(ws, 'CQ37', 'DA37'),
            "財政力指数":             _yr2(ws, 'CQ38', 'DA38'),
            "実質収支比率_percent":   _yr2(ws, 'CQ39', 'DA39'),
            "公債費負担比率_percent": _yr2(ws, 'CQ40', 'DA40'),
            "健全化判断比率": {
                "実質赤字比率_percent":     _yr2(ws, 'CQ41', 'DA41'),
                "連結実質赤字比率_percent": _yr2(ws, 'CQ42', 'DA42'),
                "実質公債費比率_percent":   _yr2(ws, 'CQ43', 'DA43'),
                "将来負担比率_percent":     _yr2(ws, 'CQ44', 'DA44'),
            },
        },
        "積立金現在高_千円": {
            "財政調整基金":         _yr2(ws, 'CQ45', 'DA45'),
            "減債基金":             _yr2(ws, 'CQ46', 'DA46'),
            "その他特定目的基金":   _yr2(ws, 'CQ47', 'DA47'),
            "定額運用基金":         _yr2(ws, 'CQ48', 'DA48'),
        },
        "土地開発基金現在高_千円": _yr2(ws, 'CQ49', 'DA49'),
        "地方債現在高_千円":       _yr2(ws, 'CQ50', 'DA50'),
        "債務負担行為額支出予定額_千円": {
            "合計":       _yr2(ws, 'CQ51', 'DA51'),
            "物件等購入": _yr2(ws, 'CQ52', 'DA52'),
            "保証・補償": _yr2(ws, 'CQ53', 'DA53'),
            "その他":     _yr2(ws, 'CQ54', 'DA54'),
            "実質的なもの": _yr2(ws, 'CQ55', 'DA55'),
        },
        "収益事業収入_千円": _yr2(ws, 'CQ56', 'DA56'),
        "事業会計の状況": {
            "国民健康保険": {
                "実質収支_千円":   _yr2(ws, 'CQ57', 'DA57'),
                "再差引収支_千円": _yr2(ws, 'CQ58', 'DA58'),
            },
        },
        "職員給与の状況": {
            "一般職員等": {
                "一般職員":         _staff(ws, 'CQ17', 'CV17', 'DB17'),
                "うち消防職員":     _staff(ws, 'CQ18', 'CV18', 'DB18'),
                "うち技能労務職員": _staff(ws, 'CQ19', 'CV19', 'DB19'),
            },
            "警察官":     _staff(ws, 'CQ20', 'CV20', 'DB20'),
            "教育公務員": _staff(ws, 'CQ21', 'CV21', 'DB21'),
            "臨時職員":   _staff(ws, 'CQ22', 'CV22', 'DB22'),
            "合計":       _staff(ws, 'CQ23', 'CV23', 'DB23'),
            "ラスパイレス指数": v(ws, 'CQ24'),
            "特別職等": {
                "知事":       _sp(ws, 'CQ27', 'CV27', 'DB27'),
                "副知事":     _sp(ws, 'CQ28', 'CV28', 'DB28'),
                "教育長":     _sp(ws, 'CQ29', 'CV29', 'DB29'),
                "議会議長":   _sp(ws, 'CQ30', 'CV30', 'DB30'),
                "議会副議長": _sp(ws, 'CQ31', 'CV31', 'DB31'),
                "議会議員":   _sp(ws, 'CQ32', 'CV32', 'DB32'),
            },
        },
    }


# ── メイン処理 ──────────────────────────────────────────────


def main(xlsx_path: Path):
    print(f"読み込み中: {xlsx_path}")
    wb = load_workbook(xlsx_path, data_only=True)
    sheets = [s for s in wb.sheetnames if s != '目次']
    print(f"対象シート数: {len(sheets)}")

    errors = []
    stack = []
    for i, name in enumerate(sheets, 1):
        try:
            data = flatten(extract_sheet(wb[name]))
            _ = json.dumps(data, ensure_ascii=False)
            stack.append(data)
            print(f"  [{i:3d}/{len(sheets)}] {name} … OK")
        except Exception as e:
            errors.append((name, str(e)))
            print(f"  [{i:3d}/{len(sheets)}] {name} … ERROR: {e}")

    print(f"\n完了: {len(stack)} 件")
    if errors:
        print(f"エラー {len(errors)} 件:")
        for name, msg in errors:
            print(f"  {name}: {msg}")

    return stack


# ── CSVに変換する処理 ──────────────────────────────────────────────
def flatten(nested: dict, sep: str = ".") -> dict:
    """Recursively flatten a nested dict."""
    out = {}

    def _flatten(obj, prefix=""):
        if obj == "-" or obj == "*":
            obj = None

        if isinstance(obj, dict):
            for k, v in obj.items():
                _flatten(v, f"{prefix}{k}{sep}" if prefix else f"{k}{sep}")
        else:
            out[prefix.rstrip(sep)] = obj

    _flatten(nested)
    return out


def add_cols(df: pl.DataFrame):
    perfmap = {}
    for idx, perf in enumerate(PERF_LIST):
        perfmap[perf] = idx

    return df.with_columns(
        pl.col('都道府県名').replace_strict(perfmap).alias('都道府県番号')).drop('都道府県名')


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(
            "Usage: python extract_pref_data.py <都道府県.xlsx> [output.msgpack]")
        sys.exit(1)

    xlsx_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2]) if len(
        sys.argv) > 2 else xlsx_path.with_name("r5_kessan_todohuken.msgpack")

    all_data = main(xlsx_path)

    stack = []
    for i in all_data:
        stack.append(i)

    df = pl.DataFrame(all_data, infer_schema_length=len(all_data))
    df = df.filter(pl.col("都道府県名").is_not_null()).drop('年度')
    df = add_cols(df)
    map_dict = df.to_dict()
    newdict = {}
    for i in map_dict:
        newdict[i] = map_dict[i].to_list()

    with open(output_path, 'wb') as fp:
        fp.write(msgpack.dumps(newdict))
        fp.close()

    print(f"出力: {output_path}")
